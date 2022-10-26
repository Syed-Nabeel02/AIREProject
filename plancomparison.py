import pandas as pd
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill
import shutil

def compare_opertional_plan_files(old_operationa_plan_directory, new_operationa_plan_directory):
    base_dir = Path(__file__).parent
    excel_1_path = old_operationa_plan_directory
    excel_2_path = new_operationa_plan_directory

    df1 = pd.read_excel(excel_1_path)
    df2 = pd.read_excel(excel_2_path)

    if df1.equals(df2):
        print("There are no changes.")
    else:
        print("There are changes.")

    wb1 = openpyxl.load_workbook(excel_1_path)
    sheet1 = wb1.active 
    wb2 = openpyxl.load_workbook(excel_2_path)
    sheet2 = wb2.active 

    excel_changes_path = base_dir / "Operational_Plan_changes.xlsx"
    shutil.copyfile(excel_1_path, excel_changes_path)
    wb = openpyxl.load_workbook(excel_changes_path)
    sheet = wb.active 
    for row in range(1, sheet.max_row):
        for col in range(1, sheet.max_column):
            if sheet1.cell(row, col).value != sheet2.cell(row, col).value:
                sheet.cell(row, col).value = '{} --> {}'.format(sheet1.cell(row, col).value, sheet2.cell(row, col).value)
                sheet.cell(row, col).fill = PatternFill(start_color="A8F3FF", fill_type = "solid")
    wb.save(excel_changes_path)

    output_dir = base_dir / "Operational_Plan_changes.xlsx"
    os.system("start " + str(output_dir))
